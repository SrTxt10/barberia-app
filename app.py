from datetime import date, datetime
from pathlib import Path
from threading import Lock

from flask import Flask, flash, redirect, render_template, request, session, url_for
from openpyxl import Workbook, load_workbook


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
EXCEL_PATH = DATA_DIR / "barberia.xlsx"

SHEETS = {
    "clientes": ["socio", "nombre", "cedula", "fecha_registro"],
    "cortes": ["id", "socio", "fecha", "servicio", "barbero", "notas"],
    "reservas": ["id", "socio", "fecha", "hora", "servicio", "estado"],
}

excel_lock = Lock()

app = Flask(__name__)
app.config["SECRET_KEY"] = "cambiar-esta-clave-en-produccion"


def ensure_database():
    DATA_DIR.mkdir(exist_ok=True)
    if EXCEL_PATH.exists():
        return

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for sheet_name, headers in SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)

    wb.save(EXCEL_PATH)


def open_workbook():
    ensure_database()
    return load_workbook(EXCEL_PATH)


def rows_as_dicts(ws):
    headers = [cell.value for cell in ws[1]]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(value is not None for value in row):
            result.append(dict(zip(headers, row)))
    return result


def next_number(ws, column=1, start=1001):
    values = []
    for cell in ws.iter_rows(min_row=2, min_col=column, max_col=column, values_only=True):
        if cell[0] is not None:
            try:
                values.append(int(cell[0]))
            except (TypeError, ValueError):
                pass
    return max(values, default=start - 1) + 1


def get_current_client():
    socio = session.get("socio")
    if not socio:
        return None

    with excel_lock:
        wb = open_workbook()
        clients = rows_as_dicts(wb["clientes"])

    for client in clients:
        if str(client["socio"]) == str(socio):
            return client
    return None


def require_login():
    client = get_current_client()
    if client:
        return client

    flash("Inicia sesion con tu numero de socio.", "warning")
    return None


def is_valid_cedula(value):
    return value.isdigit() and len(value) <= 8


@app.route("/")
def index():
    if session.get("socio"):
        return redirect(url_for("panel"))
    return render_template("index.html")


@app.route("/registro", methods=["GET", "POST"])
def registro():
    if request.method == "POST":
        nombre = request.form.get("nombre", "").strip()
        cedula = request.form.get("cedula", "").strip()

        if not nombre or not cedula:
            flash("Completa nombre y cedula.", "danger")
            return redirect(url_for("registro"))

        if not is_valid_cedula(cedula):
            flash("La cedula debe tener solo numeros y un maximo de 8 caracteres.", "danger")
            return redirect(url_for("registro"))

        with excel_lock:
            wb = open_workbook()
            ws = wb["clientes"]
            clients = rows_as_dicts(ws)

            if any(str(client["cedula"]) == cedula for client in clients):
                flash("Ya existe un cliente con esa cedula.", "danger")
                return redirect(url_for("registro"))

            socio = next_number(ws)
            ws.append([socio, nombre, cedula, date.today().isoformat()])
            wb.save(EXCEL_PATH)

        session["socio"] = str(socio)
        flash(f"Cliente registrado. Numero de socio: {socio}", "success")
        return redirect(url_for("panel"))

    return render_template("registro.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        socio = request.form.get("socio", "").strip()

        with excel_lock:
            wb = open_workbook()
            clients = rows_as_dicts(wb["clientes"])

        if any(str(client["socio"]) == socio for client in clients):
            session["socio"] = socio
            flash("Sesion iniciada.", "success")
            return redirect(url_for("panel"))

        flash("No se encontro ese numero de socio.", "danger")
        return redirect(url_for("login"))

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("Sesion cerrada.", "info")
    return redirect(url_for("index"))


@app.route("/panel")
def panel():
    client = require_login()
    if not client:
        return redirect(url_for("login"))

    socio = str(client["socio"])
    with excel_lock:
        wb = open_workbook()
        cortes = [row for row in rows_as_dicts(wb["cortes"]) if str(row["socio"]) == socio]
        reservas = [row for row in rows_as_dicts(wb["reservas"]) if str(row["socio"]) == socio]

    cortes.sort(key=lambda row: str(row["fecha"]), reverse=True)
    reservas.sort(key=lambda row: (str(row["fecha"]), str(row["hora"])))

    return render_template("panel.html", client=client, cortes=cortes, reservas=reservas)


@app.route("/reservas/nueva", methods=["GET", "POST"])
def nueva_reserva():
    client = require_login()
    if not client:
        return redirect(url_for("login"))

    if request.method == "POST":
        fecha = request.form.get("fecha", "").strip()
        hora = request.form.get("hora", "").strip()
        servicio = request.form.get("servicio", "").strip()

        if not fecha or not hora or not servicio:
            flash("Completa fecha, hora y servicio.", "danger")
            return redirect(url_for("nueva_reserva"))

        with excel_lock:
            wb = open_workbook()
            ws = wb["reservas"]
            reserva_id = next_number(ws, start=1)
            ws.append([reserva_id, client["socio"], fecha, hora, servicio, "Pendiente"])
            wb.save(EXCEL_PATH)

        flash("Reserva creada.", "success")
        return redirect(url_for("panel"))

    today = date.today().isoformat()
    return render_template("nueva_reserva.html", today=today)


@app.route("/cortes/nuevo", methods=["GET", "POST"])
def nuevo_corte():
    client = require_login()
    if not client:
        return redirect(url_for("login"))

    if request.method == "POST":
        fecha = request.form.get("fecha", "").strip() or date.today().isoformat()
        servicio = request.form.get("servicio", "").strip()
        barbero = request.form.get("barbero", "").strip()
        notas = request.form.get("notas", "").strip()

        if not servicio:
            flash("Indica el servicio realizado.", "danger")
            return redirect(url_for("nuevo_corte"))

        with excel_lock:
            wb = open_workbook()
            ws = wb["cortes"]
            corte_id = next_number(ws, start=1)
            ws.append([corte_id, client["socio"], fecha, servicio, barbero, notas])
            wb.save(EXCEL_PATH)

        flash("Corte agregado al historial.", "success")
        return redirect(url_for("panel"))

    return render_template("nuevo_corte.html", today=date.today().isoformat())


@app.template_filter("pretty_date")
def pretty_date(value):
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value)


if __name__ == "__main__":
    ensure_database()
    app.run(debug=True)
